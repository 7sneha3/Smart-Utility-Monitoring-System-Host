from flask import Blueprint, request, jsonify, send_file
from database.mongodb import utility_collection

from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
import statistics

from services.dashboard_service import (
    calculate_anomalies_from_records
)

report_bp = Blueprint(
    "report",
    __name__
    )

def get_report_type(days):
    if days == 1:
        return "Daily"

    elif days <= 7:
        return "Weekly"
    
    elif days <= 31:
        return "Monthly"

    elif days <= 90:
        return "Quarterly"

    return "Yearly"

@report_bp.route(
"/report-summary",
methods=["GET"]
)

def report_summary():

    resource_type = request.args.get(
        "type",
        "all"
    )

    sub_utility = request.args.get(
        "sub_utility",
        "all"
    )


    from_date = request.args.get(
        "from_date"
    )

    to_date = request.args.get(
        "to_date"
    )

    if not from_date or not to_date:

        return jsonify({

            "error":
            "Please select a valid date range."

        }), 400

    print("FROM =", from_date)
    print("TO =", to_date)
    query = {}

    if resource_type.lower() != "all":
        query["resource_type"] = resource_type

    if sub_utility.lower() != "all":
        query["sub_utility"] = sub_utility

    historical_records = list(
        utility_collection.find(
            query,
            {"_id": 0}
        )
    )

    records = historical_records

    if from_date and to_date:

        records = [
            r for r in historical_records
            if from_date <= r["date"] <= to_date
        ]

    if len(records) == 0:
        return jsonify({
            "records": 0,
            "total_consumption": 0,
            "avg_consumption": 0,
            "peak_consumption": 0,
            "threshold": 0,
            "anomalies": 0,
            "resource_type": resource_type,
            "sub_utility": sub_utility,
            "from_date": from_date,
            "to_date": to_date,
            "report_type": "No Data",
            "message":
                "No records found for selected date range"
        })

    consumptions = [
        r["consumption"]
        for r in records
    ]

    print(consumptions[:20])

    for i, c in enumerate(consumptions):
        if isinstance(c, str):

            print(
                "STRING FOUND AT",
                i,
                c
            )

    total = sum(consumptions)

    avg = total / len(consumptions)

    peak = max(consumptions)

    anomaly = calculate_anomalies_from_records(
        records
    )

    threshold = anomaly["threshold"]

    anomalies = len(
        anomaly["anomalies"]
    )

    print("REPORT THRESHOLD =", threshold)
    print("REPORT ANOMALIES =", anomalies)
    print("REPORT RECORDS =", len(records))
    print("THRESHOLD =", threshold)
    # print("ANOMALIES =", len(anomaly_rows))
    print("HISTORICAL =", len(historical_records))
    print("REPORT =", len(records))

    if len(records) == 0:

        return jsonify({
            "records": 0
        })

    # if not from_date:

    #     from_date = min(
    #         r["date"]
    #         for r in records
    #     )

    # if not to_date:

    #     to_date = max(
    #         r["date"]
    #         for r in records
    #     )

    days = (
        datetime.strptime(
            to_date,
            "%Y-%m-%d"
        )
        -
        datetime.strptime(
            from_date,
            "%Y-%m-%d"
        )
    ).days + 1

    report_type = get_report_type(
        days
    )

    return jsonify({

        "report_type":
            report_type,

        "records":
            len(records),

        "total_consumption":
            round(total, 2),

        "avg_consumption":
            round(avg, 2),

        "peak_consumption":
            round(peak, 2),

        "threshold":
            round(threshold, 2),

        "anomalies":
            anomalies,

        "resource_type":
            resource_type,

        "sub_utility":
            sub_utility,

        "from_date":
            from_date,

        "to_date":
            to_date
    })

@report_bp.route(
"/download-report",
methods=["GET"]
)

def download_report():
    resource_type = request.args.get(
        "type",
        "all"
    )

    sub_utility = request.args.get(
        "sub_utility",
        "all"
    )

    from_date = request.args.get(
        "from_date"
    )

    to_date = request.args.get(
        "to_date"
    )

    if not from_date or not to_date:

        return jsonify({

            "error":
            "Please select a valid date range."

        }), 400

    query = {}

    if resource_type.lower() != "all":
        query["resource_type"] = resource_type

    if sub_utility.lower() != "all":
        query["sub_utility"] = sub_utility

    records = list(
        utility_collection.find(
            query,
            {"_id": 0}
        )
    )

    if from_date and to_date:

        records = [
            r for r in records
            if from_date <= r["date"] <= to_date
        ]

    energy_total = 0
    water_total = 0

    for r in records:

        value = float(
            r["consumption"]
        )

        if (
            r["resource_type"]
            == "energy"
        ):
            energy_total += value

        elif (
            r["resource_type"]
            == "water"
        ):
            water_total += value


    if len(records) == 0:

        return jsonify({
            "error":
                "No data found"
        }), 404

    consumptions = [
        r["consumption"]
        for r in records
    ]

    print(consumptions[:20])

    for i, c in enumerate(consumptions):

        if isinstance(c, str):

            print(
                "STRING FOUND AT",
                i,
                c
            )

    total = sum(consumptions)

    avg = total / len(consumptions)

    peak = max(consumptions)

    anomaly = calculate_anomalies_from_records(
        records
    )

    threshold = anomaly["threshold"]

    anomaly_rows = anomaly["anomalies"]

    print("REPORT THRESHOLD =", threshold)
    print("REPORT ANOMALIES =", anomaly_rows)
    print("REPORT RECORDS =", len(records))
    print("THRESHOLD =", threshold)
    print("ANOMALIES =", len(anomaly_rows))

    if not from_date:

        from_date = min(
            r["date"]
            for r in records
        )

    if not to_date:

        to_date = max(
            r["date"]
            for r in records
        )

    days = (
        datetime.strptime(
            to_date,
            "%Y-%m-%d"
        )
        -
        datetime.strptime(
            from_date,
            "%Y-%m-%d"
        )
    ).days + 1

    report_type = get_report_type(
        days
    )

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Summary"

    ws1.append(
        ["Metric", "Value"]
    )

    ws1.append(
        ["Resource Type", resource_type]
    )

    ws1.append(
        ["Sub Utility", sub_utility]
    )

    ws1.append(
        ["From Date", from_date]
    )

    ws1.append(
        ["To Date", to_date]
    )

    ws1.append(
        ["Report Type", report_type]
    )

    ws1.append(
        ["Records", len(records)]
    )

    if resource_type == "all":

        ws1.append([
            "Energy Consumption",
            round(
                energy_total,
                2
            ),
            "kWh"
        ])

        ws1.append([
            "Water Consumption",
            round(
                water_total,
                2
            ),
            "m3"
        ])

    else:

        unit = ""

        if records:
            unit = records[0].get(
                "unit",
                ""
            )

        if unit == "m³":
            unit = "m3"

        ws1.append([
            "Total Consumption",
            round(
                total,
                2
            ),
            unit
        ])

    ws1.append(
        ["Average Consumption", round(avg, 2)]
    )

    ws1.append(
        ["Peak Consumption", round(peak, 2)]
    )

    ws1.append(
        ["Threshold", round(threshold, 2)]
    )

    ws1.append(
        ["Anomalies", len(anomaly_rows)]
    )

    ws2 = wb.create_sheet(
        "Consumption Data"
    )

    ws2.append([
        "Date",
        "Utility ID",
        "Sub Utility",
        "Resource Type",
        "Consumption",
        # "Unit"
    ])

    for r in records:

        unit = str(
            r.get("unit", "")
        )

        if unit == "m³":
            unit = "m3"

        ws2.append([
            r["date"],
            r["utility_id"],
            r["sub_utility"],
            r["resource_type"],
            r["consumption"],
            # unit
        ])

    ws3 = wb.create_sheet(
        "Anomaly Report"
    )

    ws3.append([
        "Date",
        "Consumption",
        "Threshold",
        "Excess"
    ])

    for row in anomaly_rows:

        ws3.append([

            row["date"],
            round(
                row["consumption"],
                2
            ),
            round(
                row["threshold"],
                2
            ),
            round(
                row["excess"],
                2
            )
        ])

    ws4 = wb.create_sheet(
        "Model Insights"
    )

    ws4.append([
        "Property",
        "Value"
    ])

    ws4.append([
        "Model Used",
        "Hybrid SVR + LSTM"
    ])

    ws4.append([
        "Features",
        "Historical Lags, Temporal Dependencies, Calendar Encoding"
    ])

    ws4.append([
        "Threshold Method",
        "Mean + 2σ"
    ])

    ws4.append([
        "Training Records",
        len(records)
    ])

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    filename = (
        f"{resource_type}_report.xlsx"
    )

    

    return send_file(

        output,

        as_attachment=True,

        download_name=filename,

        mimetype=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
